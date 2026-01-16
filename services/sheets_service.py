"""
Google Sheets integration for reading prompts and writing status
"""
import os
import json
from typing import List, Dict, Any, Optional
from pathlib import Path

try:
    import gspread
    from google.oauth2.service_account import Credentials
    GSPREAD_AVAILABLE = True
except ImportError:
    GSPREAD_AVAILABLE = False
    
from openpyxl import load_workbook, Workbook


class SheetRow:
    """Represents a row from the spreadsheet"""
    
    def __init__(
        self,
        row_index: int,
        stt: str = "",
        prompt: str = "",
        image_paths: List[str] = None,
        save_name: str = "",
        output_path: str = "",
        presets: str = "",
        status: str = "",
        type: str = "",
        aspect_ratio: str = "",  # Will be set based on type at generation time
        duration: str = "",  # Only used for video type
        resolution: str = "",  # Only used for video type
        variations: Optional[int] = None,
        model: str = "",
        extra: Dict[str, Any] = None
    ):
        self.row_index = row_index
        self.stt = stt
        self.prompt = prompt
        self.image_paths = image_paths or []
        self.save_name = save_name
        self.output_path = output_path
        self.presets = presets
        self.status = status
        self.type = type.lower() if type else ""
        self.aspect_ratio = aspect_ratio
        self.duration = duration
        self.resolution = resolution
        self.variations = variations
        self.model = model
        self.extra = extra or {}
        
    def to_dict(self) -> Dict[str, Any]:
        return {
            "row_index": self.row_index,
            "stt": self.stt,
            "prompt": self.prompt,
            "image_paths": self.image_paths,
            "save_name": self.save_name,
            "output_path": self.output_path,
            "presets": self.presets,
            "status": self.status,
            "type": self.type,
            "aspect_ratio": self.aspect_ratio,
            "duration": self.duration,
            "resolution": self.resolution,
            "variations": self.variations,
            "model": self.model,
            **self.extra
        }


class GoogleSheetsService:
    """Service for reading from Google Sheets"""
    
    # Column mappings (case-insensitive)
    COLUMN_MAPPINGS = {
        "stt": ["stt", "no", "index", "#"],
        "prompt": ["prompt", "text", "description", "nội dung"],
        "image_path": ["image", "image_path", "ref_image", "ảnh"],
        "save_name": ["savename", "save_name", "filename", "tên file"],
        "output_path": ["path", "output_path", "folder", "thư mục"],
        "presets": ["presets", "preset", "settings", "cài đặt"],
        "status": ["status", "trạng thái", "state"],
        "type": ["type", "media_type", "loại", "media"],
        "aspect_ratio": ["aspect", "aspect_ratio", "ratio", "tỷ lệ", "size"],
        "duration": ["duration", "time", "length", "thời lượng"],
        "resolution": ["resolution", "quality", "res", "độ phân giải"],
        "variations": ["variations", "variation", "variants", "variant", "num_videos", "số video", "số ảnh", "count", "số lượng"],
        "model": ["model", "style", "kiểu"],
    }
    
    def __init__(
        self,
        credentials_path: str = None,
        log_callback = None
    ):
        self.credentials_path = credentials_path
        self.log = log_callback or print
        self.client = None
        self.sheet = None
        
    def connect(self, spreadsheet_url: str = None, spreadsheet_id: str = None) -> bool:
        """Connect to Google Sheets"""
        if not GSPREAD_AVAILABLE:
            self.log("❌ gspread not installed. Use local Excel instead.")
            return False
            
        try:
            scopes = [
                "https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive"
            ]
            
            if self.credentials_path and os.path.exists(self.credentials_path):
                creds = Credentials.from_service_account_file(
                    self.credentials_path, 
                    scopes=scopes
                )
                self.client = gspread.authorize(creds)
            else:
                # Try default credentials
                self.client = gspread.service_account()
                
            # Open spreadsheet
            if spreadsheet_url:
                self.sheet = self.client.open_by_url(spreadsheet_url)
            elif spreadsheet_id:
                self.sheet = self.client.open_by_key(spreadsheet_id)
            else:
                self.log("❌ No spreadsheet URL or ID provided")
                return False
                
            self.log(f"✅ Connected to: {self.sheet.title}")
            return True
            
        except Exception as e:
            self.log(f"❌ Failed to connect: {e}")
            return False
            
    def _map_headers(self, headers: List[str]) -> Dict[str, int]:
        """Map column headers to indices"""
        header_map = {}
        headers_lower = [h.lower().strip() for h in headers]
        
        for field, aliases in self.COLUMN_MAPPINGS.items():
            for i, header in enumerate(headers_lower):
                if header in aliases:
                    header_map[field] = i
                    break
                    
        return header_map
        
    def read_worksheet(
        self,
        worksheet_name: str = None,
        worksheet_index: int = 0,
        skip_completed: bool = True
    ) -> List[SheetRow]:
        """Read all rows from worksheet"""
        if not self.sheet:
            self.log("❌ Not connected to any spreadsheet")
            return []
            
        try:
            # Get worksheet
            if worksheet_name:
                ws = self.sheet.worksheet(worksheet_name)
            else:
                ws = self.sheet.get_worksheet(worksheet_index)
                
            # Get all values
            all_values = ws.get_all_values()
            if not all_values:
                return []
                
            # Map headers
            headers = all_values[0]
            header_map = self._map_headers(headers)
            
            rows = []
            for i, row_data in enumerate(all_values[1:], start=2):  # Start from row 2
                status = row_data[header_map.get("status", -1)] if "status" in header_map else ""
                
                # Skip completed rows
                if skip_completed and status.lower() in ["done", "completed", "hoàn thành", "✅"]:
                    continue
                    
                # Parse image paths
                raw_image_val = row_data[header_map.get("image_path", -1)] if "image_path" in header_map else ""
                image_paths = [p.strip() for p in raw_image_val.split(',') if p.strip()] if raw_image_val else []

                row = SheetRow(
                    row_index=i,
                    stt=row_data[header_map.get("stt", 0)] if "stt" in header_map else str(i-1),
                    prompt=row_data[header_map.get("prompt", -1)] if "prompt" in header_map else "",
                    image_paths=image_paths,
                    save_name=row_data[header_map.get("save_name", -1)] if "save_name" in header_map else "",
                    output_path=row_data[header_map.get("output_path", -1)] if "output_path" in header_map else "",
                    presets=row_data[header_map.get("presets", -1)] if "presets" in header_map else "",
                    status=status,
                    type=row_data[header_map.get("type", -1)] if "type" in header_map and header_map.get("type", -1) >= 0 and header_map.get("type", -1) < len(row_data) else "",
                    aspect_ratio=row_data[header_map.get("aspect_ratio", -1)] if "aspect_ratio" in header_map else "16:9",
                    duration=row_data[header_map.get("duration", -1)] if "duration" in header_map else "5s",
                    model=row_data[header_map.get("model", -1)] if "model" in header_map else "",
                )
                
                # Skip empty prompts
                if row.prompt.strip():
                    rows.append(row)
                    
            self.log(f"📊 Loaded {len(rows)} rows from worksheet")
            return rows
            
        except Exception as e:
            self.log(f"❌ Failed to read worksheet: {e}")
            return []
            
    def update_status(
        self, 
        row_index: int, 
        status: str, 
        worksheet_name: str = None,
        worksheet_index: int = 0,
        status_column: str = "STATUS"
    ):
        """Update status for a specific row"""
        if not self.sheet:
            return False
            
        try:
            if worksheet_name:
                ws = self.sheet.worksheet(worksheet_name)
            else:
                ws = self.sheet.get_worksheet(worksheet_index)
                
            # Find status column
            headers = ws.row_values(1)
            status_col = None
            for i, h in enumerate(headers, start=1):
                if h.upper() == status_column.upper():
                    status_col = i
                    break
                    
            if status_col:
                ws.update_cell(row_index, status_col, status)
                self.log(f"📝 Updated row {row_index} status: {status}")
                return True
            else:
                self.log(f"⚠️ Status column not found")
                return False
                
        except Exception as e:
            self.log(f"❌ Failed to update status: {e}")
            return False


class ExcelService:
    """Service for reading from local Excel files"""
    
    COLUMN_MAPPINGS = GoogleSheetsService.COLUMN_MAPPINGS
    
    def __init__(self, log_callback = None, image_dir: str = None, output_dir: str = None):
        self.log = log_callback or print
        self.workbook = None
        self.filepath = None
        self.image_dir = Path(image_dir) if image_dir else None
        self.output_dir = Path(output_dir) if output_dir else None
        
    def load(self, filepath: str) -> bool:
        """Load Excel file"""
        try:
            self.filepath = filepath
            self.workbook = load_workbook(filepath)
            self.log(f"✅ Loaded: {filepath}")
            return True
        except Exception as e:
            self.log(f"❌ Failed to load Excel: {e}")
            return False
            
    def _map_headers(self, headers: List[str]) -> Dict[str, int]:
        """Map column headers to indices"""
        header_map = {}
        headers_lower = [str(h).lower().strip() if h else "" for h in headers]
        
        for field, aliases in self.COLUMN_MAPPINGS.items():
            for i, header in enumerate(headers_lower):
                if header in aliases:
                    header_map[field] = i
                    break
                    
        return header_map
        
    def read_worksheet(
        self,
        worksheet_name: str = None,
        skip_completed: bool = True
    ) -> List[SheetRow]:
        """Read all rows from worksheet"""
        if not self.workbook:
            self.log("❌ No workbook loaded")
            return []
            
        try:
            # Get worksheet
            if worksheet_name:
                ws = self.workbook[worksheet_name]
            else:
                ws = self.workbook.active
                
            # Get headers from first row
            headers = [cell.value for cell in ws[1]]
            header_map = self._map_headers(headers)
            
            rows = []
            for i, row_cells in enumerate(ws.iter_rows(min_row=2), start=2):
                row_data = [cell.value or "" for cell in row_cells]
                
                # Safe status extraction
                status = ""
                if "status" in header_map:
                    status_idx = header_map["status"]
                    if status_idx >= 0 and status_idx < len(row_data):
                        status = str(row_data[status_idx]) if row_data[status_idx] else ""
                
                # Skip completed rows
                if skip_completed and status.lower() in ["done", "completed", "hoàn thành", "✅"]:
                    continue
                
                def get_val(field):
                    idx = header_map.get(field, -1)
                    if idx >= 0 and idx < len(row_data):
                        val = row_data[idx]
                        return str(val).strip() if val else ""
                    return ""
                
                # Get base values
                image_name = get_val("image_path")
                save_name = get_val("save_name")
                # Note: output_path column không còn được sử dụng, chỉ dùng save_name + output_dir từ settings
                
                # Resolve image path if image_dir is set
                image_paths = []
                if image_name:
                    if self.image_dir:
                        image_paths = self._find_images(image_name)
                    else:
                        # Split by comma if multiple images provided but no dir set (raw paths)
                        image_paths = [p.strip() for p in image_name.split(',') if p.strip()]
                
                # Resolve output path - đơn giản hóa: chỉ dùng save_name và output_dir từ settings
                if save_name:
                    # Có save_name: dùng output_dir từ settings + save_name
                    if self.output_dir:
                        output_dir = self.output_dir
                    else:
                        from config.settings import DOWNLOADS_DIR
                        output_dir = Path(DOWNLOADS_DIR)
                    
                    output_dir.mkdir(parents=True, exist_ok=True)
                    save_name_with_ext = save_name if save_name.lower().endswith('.mp4') else f"{save_name}.mp4"
                    output_path = str(output_dir / save_name_with_ext)
                else:
                    # Không có save_name: tạo tên file tự động
                    if self.output_dir:
                        output_dir = self.output_dir
                    else:
                        from config.settings import DOWNLOADS_DIR
                        output_dir = Path(DOWNLOADS_DIR)
                    
                    output_dir.mkdir(parents=True, exist_ok=True)
                    output_path = str(output_dir / f"video_{get_val('stt') or i-1}.mp4")
                
                row = SheetRow(
                    row_index=i,
                    stt=get_val("stt") or str(i-1),
                    prompt=get_val("prompt"),
                    image_paths=image_paths,
                    save_name=save_name,
                    output_path=output_path,
                    presets=get_val("presets"),
                    status=status,
                    type=get_val("type") or "",
                    aspect_ratio=get_val("aspect_ratio") or "",  # No forced default
                    duration=get_val("duration") or "",  # Empty for image type
                    resolution=get_val("resolution") or "",  # Empty for image type
                    variations=int(get_val("variations")) if get_val("variations").isdigit() else None,
                    model=get_val("model"),
                )
                
                # Skip empty prompts
                if row.prompt.strip():
                    rows.append(row)
                    
            self.log(f"📊 Loaded {len(rows)} rows from Excel")
            return rows
            
        except Exception as e:
            self.log(f"❌ Failed to read worksheet: {e}")
            return []
            
    def update_status(self, row_index: int, status: str, worksheet_name: str = None):
        """Update status for a specific row"""
        if not self.workbook or not self.filepath:
            return False
            
        try:
            if worksheet_name:
                ws = self.workbook[worksheet_name]
            else:
                ws = self.workbook.active
                
            # Find status column
            headers = [cell.value for cell in ws[1]]
            status_col = None
            for i, h in enumerate(headers, start=1):
                if str(h).upper() in ["STATUS", "TRẠNG THÁI"]:
                    status_col = i
                    break
                    
            if status_col:
                ws.cell(row=row_index, column=status_col, value=status)
                self.workbook.save(self.filepath)
                self.log(f"📝 Updated row {row_index} status: {status}")
                return True
            else:
                self.log(f"⚠️ Status column not found")
                return False
                
        except Exception as e:
            self.log(f"❌ Failed to update status: {e}")
            return False
            
    def save(self):
        """Save changes to file"""
        if self.workbook and self.filepath:
            self.workbook.save(self.filepath)
            self.log(f"💾 Saved: {self.filepath}")
    
    def _find_images(self, image_names_str: str) -> List[str]:
        """Find image files by names (comma separated)"""
        if not image_names_str:
            return []
            
        found_paths = []
        image_names = [n.strip() for n in image_names_str.split(',') if n.strip()]
        
        for image_name in image_names:
            if not self.image_dir or not self.image_dir.exists():
                found_paths.append(image_name)
                continue
            
            # Common image extensions
            extensions = ['.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp']
            
            # Check if already has extension
            image_path = self.image_dir / image_name
            if image_path.exists():
                found_paths.append(str(image_path))
                continue
                
            # Try adding extensions
            found = False
            for ext in extensions:
                image_path = self.image_dir / f"{image_name}{ext}"
                if image_path.exists():
                    found_paths.append(str(image_path))
                    found = True
                    break
            
            if not found:
                # Return original if not found
                found_paths.append(image_name)
                
        return found_paths


def create_template_excel(filepath: str):
    """Create a template Excel file with all supported columns"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Prompts"
    
    # Headers - All supported columns
    headers = [
        "STT",           # Row number / index
        "PROMPT",        # The prompt text for generation (REQUIRED)
        "IMAGE",         # Reference image file name(s), comma-separated
        "SAVENAME",      # Output file/folder name (without extension)
        "TYPE",          # "image" or "video" (default: image)
        "ASPECT_RATIO",  # "16:9", "9:16", "1:1", "3:2", "2:3" (default: 16:9)
        "DURATION",      # "5s", "10s", "15s", "20s" - only for video (default: 10s)
        "RESOLUTION",    # "480p", "720p", "1080p" - only for video (default: 480p)
        "VARIATIONS",    # 1, 2, or 4 - number of outputs (default: 1)
        "STATUS"         # Auto-updated by tool: Completed, Failed, Processing
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        
    # Sample rows showing different use cases
    sample_data = [
        # Row 1: Simple video generation
        ["1", "A cinematic shot of a sunrise over mountains with golden rays breaking through clouds", "", "sunrise_video", "video", "16:9", "10s", "720p", "1", ""],
        # Row 2: Image generation with custom aspect ratio
        ["2", "Close-up of coffee being poured into a ceramic cup, steam rising, warm lighting", "", "coffee_pour", "image", "1:1", "", "", "2", ""],
        # Row 3: Video from reference image
        ["3", "A cat walking gracefully across a moonlit garden", "cat_ref.jpg", "moonlit_cat", "video", "9:16", "5s", "480p", "1", ""],
        # Row 4: Multiple image variations
        ["4", "A futuristic cityscape with flying cars and neon lights at night", "", "cyber_city", "image", "16:9", "", "", "4", ""],
    ]
    
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Add a description sheet
    desc_ws = wb.create_sheet("Instructions")
    instructions = [
        ["COLUMN", "DESCRIPTION", "REQUIRED", "VALUES"],
        ["STT", "Row number (auto-generated if empty)", "No", "1, 2, 3..."],
        ["PROMPT", "Description of what to generate", "YES", "Any text"],
        ["IMAGE", "Reference image file name(s)", "No", "image.jpg or img1.jpg,img2.jpg"],
        ["SAVENAME", "Output file name (no extension)", "No", "my_video (creates: my_video/my_video_01.mp4)"],
        ["TYPE", "Media type to generate", "No", "image (default) or video"],
        ["ASPECT_RATIO", "Output aspect ratio", "No", "16:9, 9:16, 1:1, 3:2, 2:3"],
        ["DURATION", "Video length (video only)", "No", "5s, 10s, 15s, 20s"],
        ["RESOLUTION", "Video quality (video only)", "No", "480p, 720p, 1080p"],
        ["VARIATIONS", "Number of outputs to generate", "No", "1, 2, or 4"],
        ["STATUS", "Auto-updated by tool", "No", "Don't edit - filled automatically"],
    ]
    for row_idx, row_data in enumerate(instructions, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            desc_ws.cell(row=row_idx, column=col_idx, value=value)
            
    wb.save(filepath)
    print(f"✅ Template created: {filepath}")

