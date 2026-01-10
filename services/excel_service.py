"""
Excel Service - Read prompts and config from Excel file
"""
import os
from pathlib import Path
from typing import List, Dict, Optional
from dataclasses import dataclass, field

try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    load_workbook = None
    Workbook = None


@dataclass
class JobRow:
    """Represents a single job from Excel"""
    stt: int = 0
    prompt: str = ""
    image: str = ""  # Image filename (without extension)
    savename: str = ""  # Output video name
    path: str = ""  # Subfolder in Output/
    presets: str = ""
    status: str = ""
    row_index: int = 0
    
    # Runtime fields
    image_path: str = ""  # Full path to image file
    output_path: str = ""  # Full path for output video
    
    def to_dict(self) -> dict:
        return {
            'stt': self.stt,
            'prompt': self.prompt,
            'image': self.image,
            'savename': self.savename,
            'path': self.path,
            'presets': self.presets,
            'status': self.status,
            'row_index': self.row_index,
            'image_path': self.image_path,
            'output_path': self.output_path,
        }


class ExcelService:
    """Service for reading/writing Excel files"""
    
    # Map column names (lowercase, no spaces) to JobRow fields
    COLUMN_MAP = {
        'stt': 'stt',
        'prompt': 'prompt',
        'image': 'image',
        'savename': 'savename',
        'save_name': 'savename',
        'path': 'path',
        'aspect_ratio': 'aspect_ratio',
        'aspectratio': 'aspect_ratio',
        'duration': 'duration',
        'presets': 'presets',
        'preset': 'presets',
        'status': 'status',
        'model': 'model',
    }
    
    def __init__(self, image_dir: str, output_dir: str, log_callback=None):
        self.image_dir = Path(image_dir)
        self.output_dir = Path(output_dir)
        self.log = log_callback or print
        self.excel_path = None
        self.workbook = None
        
    def load_excel(self, filepath: str) -> List[JobRow]:
        """Load jobs from Excel file"""
        if load_workbook is None:
            self.log("❌ openpyxl chưa cài. Chạy: pip install openpyxl")
            return []
            
        filepath = Path(filepath)
        if not filepath.exists():
            # Try adding .xlsx extension
            if not filepath.suffix:
                filepath = filepath.with_suffix('.xlsx')
            if not filepath.exists():
                self.log(f"❌ File không tồn tại: {filepath}")
                return []
                
        self.excel_path = filepath
        
        try:
            self.workbook = load_workbook(filepath)
            ws = self.workbook.active
            
            # Get column headers
            headers = {}
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value:
                    key = str(cell.value).strip().lower().replace(' ', '_')
                    headers[col_idx] = key
                    
            self.log(f"📋 Columns: {list(headers.values())}")
            
            # Parse rows
            jobs = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                job = JobRow(row_index=row_idx)
                
                for col_idx, cell in enumerate(row, 1):
                    if col_idx not in headers:
                        continue
                    col_name = headers[col_idx]
                    field_name = self.COLUMN_MAP.get(col_name)
                    
                    if field_name and hasattr(job, field_name):
                        value = cell.value
                        if value is not None:
                            if field_name == 'stt':
                                job.stt = int(value) if value else 0
                            else:
                                setattr(job, field_name, str(value).strip())
                                
                # Skip rows without prompt
                if not job.prompt:
                    continue
                    
                # Resolve image path
                if job.image:
                    job.image_path = self._find_image(job.image)
                    
                # Resolve output path
                job.output_path = self._get_output_path(job)
                
                jobs.append(job)
                
            self.log(f"✅ Loaded {len(jobs)} jobs từ Excel")
            return jobs
            
        except Exception as e:
            self.log(f"❌ Lỗi đọc Excel: {e}")
            return []
            
    def _find_image(self, image_name: str) -> str:
        """Find image file by name (with or without extension)"""
        # Common image extensions
        extensions = ['.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp']
        
        # Check if already has extension
        image_path = self.image_dir / image_name
        if image_path.exists():
            return str(image_path)
            
        # Try adding extensions
        for ext in extensions:
            image_path = self.image_dir / f"{image_name}{ext}"
            if image_path.exists():
                return str(image_path)
                
        return ""
        
    def _get_output_path(self, job: JobRow) -> str:
        """Get output path for video"""
        # Create subfolder based on PATH column
        if job.path:
            output_folder = self.output_dir / job.path
        else:
            output_folder = self.output_dir
            
        output_folder.mkdir(parents=True, exist_ok=True)
        
        # Generate filename
        if job.savename:
            filename = f"{job.savename}.mp4"
        else:
            filename = f"video_{job.stt}.mp4"
            
        return str(output_folder / filename)
        
    def update_status(self, row_index: int, status: str) -> bool:
        """Update STATUS column for a row"""
        if not self.workbook or not self.excel_path:
            return False
            
        try:
            ws = self.workbook.active
            
            # Find STATUS column
            status_col = None
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value and str(cell.value).strip().upper() == 'STATUS':
                    status_col = col_idx
                    break
                    
            if status_col:
                ws.cell(row=row_index, column=status_col, value=status)
                self.workbook.save(self.excel_path)
                return True
                
        except Exception as e:
            self.log(f"⚠️ Không update được STATUS: {e}")
            
        return False
        
    @staticmethod
    def create_template(filepath: str):
        """Create Excel template"""
        if Workbook is None:
            return False
            
        wb = Workbook()
        ws = wb.active
        ws.title = "Prompts"
        
        # Headers
        headers = ['STT', 'PROMPT', 'IMAGE', 'SAVENAME', 'PATH', 'PRESETS', 'STATUS']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            
        # Sample data
        ws.cell(row=2, column=1, value=1)
        ws.cell(row=2, column=2, value="A cinematic video of a sunset over mountains")
        ws.cell(row=2, column=3, value="sample")
        ws.cell(row=2, column=4, value="sunset_video")
        ws.cell(row=2, column=5, value="Nature")
        
        wb.save(filepath)
        return True


def test_service():
    """Test the service"""
    service = ExcelService(
        image_dir="C:/Projects/Tool Auto Veo AI/python-tool-sora/Image",
        output_dir="C:/Projects/Tool Auto Veo AI/python-tool-sora/Output"
    )
    
    jobs = service.load_excel("C:/Projects/Tool Auto Veo AI/Book1.xlsx")
    for job in jobs:
        print(f"Job {job.stt}: {job.prompt[:50]}...")
        print(f"  Image: {job.image_path}")
        print(f"  Output: {job.output_path}")
        print()


if __name__ == "__main__":
    test_service()
